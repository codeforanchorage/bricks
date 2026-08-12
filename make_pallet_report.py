#!/usr/bin/env python3
"""Build the warehouse Excel workbook: one tab per PALLET, not per section.

make_report.py answers "does this official brick exist?" and is organised
by the park section printed in the official list. This workbook answers
the warehouse question -- "what is actually stacked on this pallet?" --
so staff pulling a brick can work from the pallet they are standing at.

Sheets:
  Summary      per-pallet totals and which sections the pallet holds
  All bricks   every row from every pallet tab in one sheet, sorted by
               ORIGINAL brick number -- the search-everything view
  Pallet <X>   one row per PHOTO taken on that pallet in original-
               brick-number order, with the brick it matched (section,
               number, inscription) or its review state

Sorting is numeric on the original number (certificates carry it;
leading zeros like 0021 sort as 21), falling back to the new number
when a row has no original one.

Every row leads with the Section and Pallet cells highlighted in the
search page's badge colours (navy section, gold pallet).

Pallet labels are warehouse labels, not sections: most pallets are ~90%
one section but every one carries strays (and 'Pallet H2' is mostly
section F), so the Section column on each tab is the retrieval truth.

Built for the printed-page use case (a PDF of these tabs posted at each
pallet): every row carries BOTH brick numbers (people arrive with either
the original number or the post-2009 one), the buyer surname from both
lists (OG list = OCR'd scan, new list = clean typed Excel), the new
list's inscription, and the photo's own OCR read for comparison.

Usage:
    python make_pallet_report.py --matched output/pallets_final.csv \
        --master reference/master_list.csv \
        --output output/brick_report_by_pallet.xlsx
"""
from __future__ import annotations

import argparse
import csv
import sys
from collections import Counter, defaultdict
from pathlib import Path

HEADERS = ["Section", "Pallet", "Orig #", "New #",
           "Buyer (OG list)", "Buyer (new list)", "Photo OCR read",
           "Inscription (new list)", "Photo", "Status", "Review note"]
WIDTHS = [9, 11, 9, 9, 20, 20, 36, 44, 20, 11, 26]
STATUS_COL = HEADERS.index("Status")

STATUS = {"matched": "Present", "unmatched": "in review",
          "no_match": "unofficial", "stack_photo": "stack shot",
          "illegible": "illegible"}


ORIG_COL = HEADERS.index("Orig #")
NEW_COL = HEADERS.index("New #")
PHOTO_COL = HEADERS.index("Photo")


def _orig_key(values: list) -> tuple:
    # Sort by ORIGINAL brick number -- certificates carry it. Leading
    # zeros are display-only: 0021 sorts as 21. Rows with no original
    # number fall back to the new number; number-less rows (in-review
    # photos, unofficial bricks) sort last, by photo name.
    for ident in (values[ORIG_COL], values[NEW_COL]):
        digits = str(ident or "").replace(",", "").strip()
        if digits.isdigit():
            return (0, int(digits), values[PHOTO_COL])
    return (1, 10**9, values[PHOTO_COL])


def main(argv=None) -> None:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--matched", required=True, type=Path,
                        help="Applied catalogue (output/pallets_final.csv)")
    parser.add_argument("--master", required=True, type=Path,
                        help="reference/master_list.csv (buyer + both "
                             "brick numbers)")
    parser.add_argument("--xls", type=Path,
                        default=Path("reference/brick_list_xls.csv"),
                        help="Official Excel list (clean typed buyer "
                             "key_word per section+new id)")
    parser.add_argument("--output", required=True, type=Path,
                        help=".xlsx workbook to write")
    args = parser.parse_args(argv)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise SystemExit("openpyxl is not installed -- run: "
                         "pip install -r requirements.txt")

    # Master rows keyed the way match.py reports a brick: (section, the
    # id it matched under -- new_id when the brick was renumbered).
    master: dict[tuple[str, str], dict] = {}
    with open(args.master, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            key = (row["section"].upper(), row["new_id"] or row["orig_id"])
            master.setdefault(key, row)  # dup_orig twins: first row speaks

    # The official Excel list's key_word is the clean TYPED buyer name --
    # the master's buyer column is the OCR'd OG-list surname.
    xls_buyer: dict[tuple[str, str], str] = {}
    if args.xls.exists():
        with open(args.xls, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                key = (row["section"].upper(), row["assigned_id"])
                xls_buyer.setdefault(key, row.get("key_word", ""))

    by_pallet: dict[str, list[dict]] = defaultdict(list)
    with open(args.matched, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            by_pallet[row.get("pallet", "") or "(no pallet)"].append(row)

    wb = Workbook()
    bold = Font(bold=True)
    green = PatternFill("solid", start_color="D8EFD0")
    # Section and Pallet lead every row highlighted in the search page's
    # badge colours: navy = the brick's official section, gold = the
    # pallet its photo came from.
    sec_fill = PatternFill("solid", start_color="274156")
    sec_font = Font(bold=True, color="FFFFFF")
    pal_fill = PatternFill("solid", start_color="FFD966")
    pal_font = Font(bold=True, color="4A3A00")

    def _paint(ws) -> None:
        cell_sec, cell_pal = ws[ws.max_row][0], ws[ws.max_row][1]
        cell_sec.fill, cell_sec.font = sec_fill, sec_font
        cell_pal.fill, cell_pal.font = pal_fill, pal_font

    summary = wb.active
    summary.title = "Summary"
    summary.append(["Pallet", "Photos", "Present", "Distinct bricks",
                    "In review", "Other", "Sections on this pallet"])
    for cell in summary[1]:
        cell.font = bold
    for i, width in enumerate((12, 8, 9, 14, 10, 7, 50), 1):
        summary.column_dimensions[get_column_letter(i)].width = width

    totals = Counter()
    all_keys: set = set()
    all_rows: list[list] = []  # every tab's sheet values
    for pallet in sorted(by_pallet):
        rows = by_pallet[pallet]
        ws = wb.create_sheet(pallet[:31])
        ws.append(HEADERS)
        for cell in ws[1]:
            cell.font = bold
        for i, width in enumerate(WIDTHS, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        ws.freeze_panes = "A2"

        sections = Counter()
        keys = set()
        n_present = n_review = n_other = 0
        prepared: list[list] = []
        for row in rows:
            status = STATUS.get(row.get("match_status", ""),
                                row.get("match_status", ""))
            if status == "Present":
                n_present += 1
                sections[row.get("official_section", "") or "?"] += 1
                keys.add((row.get("official_section", ""),
                          row.get("official_id", "")))
            elif status == "in review":
                n_review += 1
            else:
                n_other += 1
            key = (row.get("official_section", "").upper(),
                   row.get("official_id", ""))
            m = master.get(key, {})
            prepared.append([
                row.get("official_section", ""),
                pallet,
                m.get("orig_id", ""),
                m.get("new_id", ""),
                m.get("buyer", ""),
                xls_buyer.get(key, ""),
                row.get("matched_read", ""),
                m.get("new_inscription", ""),
                Path(row.get("image", "")).name,
                status,
                row.get("review_note", ""),
            ])
        prepared.sort(key=_orig_key)
        for values in prepared:
            ws.append(values)
            all_rows.append(values)
            if values[STATUS_COL] == "Present":
                for cell in ws[ws.max_row]:
                    cell.fill = green
            _paint(ws)
        ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"

        section_note = ", ".join(f"{s}: {n}" for s, n in sections.most_common())
        summary.append([pallet, len(rows), n_present, len(keys),
                        n_review, n_other, section_note])
        totals.update({"photos": len(rows), "present": n_present,
                       "review": n_review, "other": n_other})
        all_keys |= keys

    # One searchable sheet with every pallet's rows, in original-number
    # order so a lookup doesn't need to know the pallet.
    all_ws = wb.create_sheet("All bricks", 1)
    all_ws.append(HEADERS)
    for cell in all_ws[1]:
        cell.font = bold
    for i, width in enumerate(WIDTHS, 1):
        all_ws.column_dimensions[get_column_letter(i)].width = width
    all_ws.freeze_panes = "A2"
    for values in sorted(all_rows, key=_orig_key):
        all_ws.append(values)
        if values[STATUS_COL] == "Present":
            for cell in all_ws[all_ws.max_row]:
                cell.fill = green
        _paint(all_ws)
    all_ws.auto_filter.ref = (f"A1:{get_column_letter(len(HEADERS))}"
                              f"{all_ws.max_row}")

    summary.append(["TOTAL", totals["photos"], totals["present"],
                    len(all_keys), totals["review"], totals["other"], ""])
    for cell in summary[summary.max_row]:
        cell.font = bold
    summary.append([])
    summary.append(["Pallet labels are warehouse labels, not sections -- "
                    "the Section column on each tab says where the brick "
                    "belongs in the official list."])
    summary.append(["People may arrive with EITHER brick number: Orig # is "
                    "from pre-2009 certificates, New # is the current "
                    "official numbering."])
    summary.append(["Buyer (OG list) is OCR'd from the scanned original "
                    "list; Buyer (new list) is the clean typed name from "
                    "the official Excel. Photo OCR read is what the camera "
                    "saw on the brick itself."])
    summary.append(["'Distinct bricks' can be below 'Present' when the same "
                    "brick was photographed twice on one pallet; the TOTAL "
                    "row de-duplicates across pallets."])

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"Wrote {args.output}")
    print(f"  {len(by_pallet)} pallet tab(s); {totals['photos']} photo(s); "
          f"{totals['present']} Present ({len(all_keys)} distinct "
          f"brick(s)); {totals['review']} in review")


if __name__ == "__main__":
    main()
