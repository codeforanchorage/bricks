#!/usr/bin/env python3
"""Build the Parks & Rec Excel workbook from the master list + photo results.

This is the handoff document: one row per official brick, annotated with
whether a warehouse photo has confirmed it. Staff at the pickup counter
filter/search it to answer "does this person's brick exist, which section,
and have we actually seen it?".

Sheets:
  Summary     per-section totals: bricks, photographed, review, remaining
  All bricks  every master row with its photo status
  A..K        the same rows split per section (what a pallet crew works from)

Photo status values:
  Present     a photo matched this brick (machine >= threshold, or human)
  (blank)     no photo of this brick yet -- NOT evidence it is missing until
              its section is photographed in full

Photo-side outcomes that never mark a brick Present but are reported:
  unmatched   still in the human review queue (counted on the Summary)
  no_match    reviewer confirmed the brick is NOT in the official lists --
              these appear on their own 'Unofficial bricks' sheet with the
              photo's text, pallet, and the reviewer's note (the photo IS
              the record for these)
  stack_photo reviewer confirmed a pallet/stack overview shot (not a brick)
  illegible   reviewer confirmed the photo is unreadable

Reviewer notes ride along: a brick's Present row carries the note its
reviewer wrote, and unofficial bricks carry theirs.

Accepts one or more matched catalogues (match.py / apply_decisions.py
output); rows whose match_status is 'matched' claim their official brick.
Pass the APPLIED catalogue (pallets_final.csv) so human decisions and
notes are included.

Usage:
    python make_report.py --master reference/master_list.csv \
        --matched output/matched_final.csv --output output/brick_report.xlsx
"""
from __future__ import annotations

import argparse
import csv
import sys
from collections import defaultdict
from pathlib import Path

from consensus import _normalise, _similar

HEADERS = ["Section", "Orig #", "New #", "Buyer", "Inscription",
           "List status", "Flag", "Photo status", "Photo", "Reviewer",
           "Review note"]
WIDTHS = [9, 9, 9, 24, 46, 11, 16, 13, 22, 14, 30]

UNOFFICIAL_HEADERS = ["Photo", "Pallet", "Brick reads (OCR)", "Reviewer",
                      "Review note"]
UNOFFICIAL_WIDTHS = [34, 12, 44, 14, 34]


def _load_matches(paths: list[Path]) -> tuple[dict, dict, list[dict]]:
    """Photo results keyed by (section, official id).

    Returns (present, counts, unofficial): present maps the key to the
    best matched photo row; counts tallies the photo-side outcomes that
    never mark a brick Present; unofficial collects human-confirmed
    no_match photos -- bricks that exist at the pickup site but are
    absent from the official lists.
    """
    present: dict[tuple[str, str], dict] = {}
    counts = {"unmatched": 0, "stack_photo": 0, "illegible": 0}
    unofficial: list[dict] = []
    for path in paths:
        with open(path, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                status = row.get("match_status", "")
                if status == "matched":
                    key = (row.get("official_section", "").upper(),
                           row.get("official_id", ""))
                    best = present.get(key)
                    if best is None or float(row.get("score") or 0) > \
                            float(best.get("score") or 0):
                        present[key] = row
                elif status == "no_match":
                    unofficial.append(row)
                elif status in counts:
                    counts[status] += 1
    return present, counts, unofficial


def _display_inscription(row: dict) -> str:
    """Human-facing text: workbook > clean re-read (0.40 rule) > parse."""
    if row["new_inscription"]:
        return row["new_inscription"]
    og, alt = row["og_inscription"], row.get("og_alt", "")
    if alt and _similar(_normalise(alt), _normalise(og)) >= 0.40:
        return alt
    return og


def main(argv=None) -> None:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--master", required=True, type=Path,
                        help="reference/master_list.csv from merge_lists.py")
    parser.add_argument("--matched", type=Path, nargs="*", default=[],
                        help="Matched catalogue(s) from match.py / "
                             "apply_decisions.py (omit for a list-only book)")
    parser.add_argument("--output", required=True, type=Path,
                        help=".xlsx workbook to write")
    args = parser.parse_args(argv)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise SystemExit("openpyxl is not installed -- run: "
                         "pip install -r requirements.txt")

    with open(args.master, newline="", encoding="utf-8") as f:
        master = list(csv.DictReader(f))
    present, counts, unofficial = _load_matches(args.matched)

    def brick_row(row: dict) -> list[str]:
        section = row["section"].upper()
        # A master row's photo key: the id match.py reports for it.
        key = (section, row["new_id"] or row["orig_id"])
        photo = present.get(key)
        return [
            section or "?",
            row["orig_id"],
            row["new_id"],
            row["buyer"],
            _display_inscription(row),
            row["status"],
            row["flag"],
            "Present" if photo else "",
            photo.get("image", "") if photo else "",
            photo.get("reviewer", "") if photo else "",
            photo.get("review_note", "") if photo else "",
        ]

    wb = Workbook()
    bold = Font(bold=True)
    green = PatternFill("solid", start_color="D8EFD0")

    def start_sheet(ws):
        ws.append(HEADERS)
        for cell in ws[1]:
            cell.font = bold
        for i, width in enumerate(WIDTHS, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        ws.freeze_panes = "A2"

    def finish_sheet(ws):
        ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"

    # --- All bricks ------------------------------------------------------
    all_ws = wb.active
    all_ws.title = "All bricks"
    start_sheet(all_ws)
    by_section: dict[str, list[dict]] = defaultdict(list)
    # Present counts distinct photographed BRICKS (keys), not master rows:
    # dup_orig twin-rows share an id, so one photo lights up both rows --
    # counting rows would inflate the Summary.
    section_stats: dict[str, dict] = defaultdict(
        lambda: {"total": 0, "present_keys": set()})
    for row in master:
        section = row["section"].upper() or "?"
        by_section[section].append(row)
        values = brick_row(row)
        all_ws.append(values)
        section_stats[section]["total"] += 1
        if values[7] == "Present":
            key = (row["section"].upper(), row["new_id"] or row["orig_id"])
            section_stats[section]["present_keys"].add(key)
            for cell in all_ws[all_ws.max_row]:
                cell.fill = green
    finish_sheet(all_ws)

    # --- Per-section sheets ----------------------------------------------
    for section in sorted(by_section):
        ws = wb.create_sheet(f"Section {section}" if section != "?"
                             else "Unassigned")
        start_sheet(ws)
        rows = sorted(by_section[section],
                      key=lambda r: int((r["new_id"] or r["orig_id"])
                                        .replace(",", ""))
                      if (r["new_id"] or r["orig_id"]).replace(",", "").isdigit()
                      else 0)
        for row in rows:
            values = brick_row(row)
            ws.append(values)
            if values[7] == "Present":
                for cell in ws[ws.max_row]:
                    cell.fill = green
        finish_sheet(ws)

    # --- Unofficial bricks (photo is the record) --------------------------
    if unofficial:
        ws = wb.create_sheet("Unofficial bricks")
        ws.append(UNOFFICIAL_HEADERS)
        for cell in ws[1]:
            cell.font = bold
        for i, width in enumerate(UNOFFICIAL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        ws.freeze_panes = "A2"
        for row in sorted(unofficial, key=lambda r: r.get("image", "")):
            ws.append([row.get("image", ""), row.get("pallet", ""),
                       row.get("matched_read", ""),
                       row.get("reviewer", ""),
                       row.get("review_note", "")])
        ws.auto_filter.ref = (f"A1:{get_column_letter(len(UNOFFICIAL_HEADERS))}"
                              f"{ws.max_row}")

    # --- Summary (inserted first) ------------------------------------------
    summary = wb.create_sheet("Summary", 0)
    summary.append(["Section", "Bricks", "Photographed (Present)",
                    "Not yet photographed"])
    for cell in summary[1]:
        cell.font = bold
    for i, width in enumerate((10, 10, 24, 22), 1):
        summary.column_dimensions[get_column_letter(i)].width = width
    total = present_total = 0
    for section in sorted(section_stats):
        stats = section_stats[section]
        present = len(stats["present_keys"])
        summary.append([section, stats["total"], present,
                        stats["total"] - present])
        total += stats["total"]
        present_total += present
    summary.append(["TOTAL", total, present_total, total - present_total])
    for cell in summary[summary.max_row]:
        cell.font = bold
    notes = []
    if counts["unmatched"]:
        notes.append(f"{counts['unmatched']} photo(s) are still in review "
                     f"and not counted as Present.")
    if unofficial:
        notes.append(f"{len(unofficial)} photo(s) show bricks that exist at "
                     f"the pickup site but are NOT in the official lists -- "
                     f"see the 'Unofficial bricks' sheet.")
    if counts["stack_photo"]:
        notes.append(f"{counts['stack_photo']} photo(s) are pallet/stack "
                     f"overview shots (not bricks).")
    if counts["illegible"]:
        notes.append(f"{counts['illegible']} photo(s) were confirmed "
                     f"unreadable.")
    for note in notes:
        summary.append([])
        summary.append([note])
    summary.append([])
    summary.append(["A blank Photo status means the brick has not been "
                    "photographed yet -- it is NOT evidence the brick is "
                    "missing until its section is photographed in full."])

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"Wrote {args.output}")
    print(f"  {total} brick(s); {present_total} confirmed present by photo; "
          f"{counts['unmatched']} photo(s) still in review; "
          f"{len(unofficial)} unofficial brick(s); "
          f"{counts['stack_photo']} stack shot(s)")


if __name__ == "__main__":
    main()
